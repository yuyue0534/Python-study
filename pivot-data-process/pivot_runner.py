"""
pivot_runner.py — 透视表自动化脚本
====================================
用法：python pivot_runner.py
依赖：pip install pandas openpyxl pyyaml
"""

import sys
import os
from pathlib import Path
from datetime import datetime

import yaml
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 路径基准：脚本所在目录 ───────────────────────────────────────
BASE_DIR = Path(__file__).parent

# ── 必须存在的原始字段（用于校验 SQL 导出是否完整）───────────────
REQUIRED_COLUMNS = [
    "SCHEME_CODE", "BILL_PERIOD", "BILL_STATUS", "BILL_REF_NO",
    "MB_CNT", "BILL_CON_AMT", "RPT_AMT", "PAID_AMT", "OS_AMT",
    # 以下字段在 SQL 导出中可能已是计算好的值，也可能需要复现
    "RS_SUBMIT", "RS_SUBMIT_MEM", "PAID", "PAID_MEM",
    # 用于 Zero_Bill 计算
    "BILL_CON_AMT",
]

# ── 透视表值字段配置（列标题 -> (源字段, 聚合函数)）────────────────
PIVOT_METRICS = [
    ("# of Pre-gen RS",       "BILL_REF_NO",     "count"),
    ("Sum of MB_CNT",         "MB_CNT",           "sum"),
    ("Bill Amount",           "BILL_CON_AMT",     "sum"),
    ("# of Submit RS",        "RS_SUBMIT",        "sum"),
    ("Sum of RS_SUBMIT_MEM",  "RS_SUBMIT_MEM",    "sum"),
    ("Submit Amount",         "RPT_AMT",          "sum"),
    ("# of Paid RS",          "PAID",             "sum"),
    ("Sum of PAID_MEM",       "PAID_MEM",         "sum"),
    ("Sum of PAID_AMT",       "PAID_AMT",         "sum"),
    ("Sum of OS_AMT",         "OS_AMT",           "sum"),
]

# ── 行分组字段（4 层层级）────────────────────────────────────────
GROUPBY_KEYS = ["SCHEME_CODE", "BILL_PERIOD", "Zero_Bill", "BILL_STATUS"]

# ── Zero_Bill 分类顺序（控制透视表行的排列顺序）─────────────────
ZERO_BILL_ORDER = ["1-No Member", "2-Zero Bill", "3-Positive Bill"]


# ════════════════════════════════════════════════════════════════
# Step 1: 读取配置
# ════════════════════════════════════════════════════════════════
def load_config() -> dict:
    config_path = BASE_DIR / "config.yaml"
    if not config_path.exists():
        _abort(f"找不到配置文件：{config_path}\n请确保 config.yaml 与脚本在同一目录。")

    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    # 解析输入路径（支持相对路径）
    input_path = Path(cfg["input_file"])
    if not input_path.is_absolute():
        input_path = BASE_DIR / input_path
    cfg["_input_path"] = input_path

    return cfg


# ════════════════════════════════════════════════════════════════
# Step 2: 读取原始数据
# ════════════════════════════════════════════════════════════════
def load_data(cfg: dict) -> pd.DataFrame:
    path = cfg["_input_path"]
    sheet = cfg.get("input_sheet", "raw")

    if not path.exists():
        _abort(
            f"找不到输入文件：{path}\n"
            f"请检查 config.yaml 中的 input_file 路径是否正确。"
        )

    _info(f"正在读取数据：{path.name}  Sheet='{sheet}' ...")

    try:
        df = pd.read_excel(path, sheet_name=sheet, dtype={"BILL_PERIOD": int})
    except Exception as e:
        _abort(f"读取文件失败：{e}")

    _info(f"读取完成：{len(df):,} 行 × {len(df.columns)} 列")

    # 校验必要列
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        _abort(
            f"原始数据缺少以下必要字段，请检查 SQL 导出是否完整：\n"
            + "\n".join(f"  · {c}" for c in missing)
        )

    return df


# ════════════════════════════════════════════════════════════════
# Step 3: 复现计算列
# ════════════════════════════════════════════════════════════════
def build_computed_cols(df: pd.DataFrame) -> pd.DataFrame:
    _info("正在复现计算列 ...")
    df = df.copy()

    mb  = df["MB_CNT"].fillna(0)
    amt = df["BILL_CON_AMT"].fillna(0)

    # Zero_Bill 分类
    conditions = [mb == 0, (mb > 0) & (amt == 0), (mb > 0) & (amt > 0)]
    df["Zero_Bill"] = np.select(conditions, ZERO_BILL_ORDER, default="1-No Member")

    # RS_SUBMIT / RS_SUBMIT_MEM / PAID / PAID_MEM
    # 若 SQL 导出已包含正确数值则跳过；若为公式字符串则重新计算
    for col in ["RS_SUBMIT", "RS_SUBMIT_MEM", "PAID", "PAID_MEM"]:
        if df[col].dtype == object:
            _warn(f"列 '{col}' 读取到的是公式文本，将根据规则重新计算。")
            if col == "RS_SUBMIT":
                # 是否有提交记录：PAY_SUBMIT_REF_NO 或 RS_SUBMIT_CNT > 0
                df[col] = _coerce_int(
                    df.get("PAY_SUBMIT_REF_NO", pd.Series()).notna() |
                    (df.get("RS_SUBMIT_CNT", pd.Series(0)).fillna(0) > 0)
                )
            elif col == "RS_SUBMIT_MEM":
                df[col] = np.where(df["RS_SUBMIT"].astype(int) > 0, mb.astype(int), 0)
            elif col == "PAID":
                df[col] = _coerce_int(df.get("PAY_CNT", pd.Series(0)).fillna(0) > 0)
            elif col == "PAID_MEM":
                df[col] = np.where(df["PAID"].astype(int) > 0, mb.astype(int), 0)

    return df


def _coerce_int(series) -> pd.Series:
    return series.astype(int) if hasattr(series, "astype") else pd.Series(0)


# ════════════════════════════════════════════════════════════════
# Step 4: 应用切片器筛选条件
# ════════════════════════════════════════════════════════════════
def apply_filters(df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    filters = cfg.get("filters", {})
    scheme  = filters.get("SCHEME_CODE", "FD")
    period  = int(filters.get("BILL_PERIOD", 202511))

    _info(f"正在筛选：SCHEME_CODE='{scheme}'  BILL_PERIOD={period} ...")

    df_f = df[(df["SCHEME_CODE"] == scheme) & (df["BILL_PERIOD"] == period)]

    if df_f.empty:
        _abort(
            f"筛选后数据为空！\n"
            f"原始数据中 SCHEME_CODE 的唯一值：{df['SCHEME_CODE'].unique().tolist()}\n"
            f"原始数据中 BILL_PERIOD 的唯一值：{sorted(df['BILL_PERIOD'].unique().tolist())}"
        )

    _info(f"筛选完成：剩余 {len(df_f):,} 行")
    return df_f


# ════════════════════════════════════════════════════════════════
# Step 5: 构建透视表
# ════════════════════════════════════════════════════════════════
def build_pivot(df: pd.DataFrame) -> pd.DataFrame:
    _info("正在构建透视表 ...")

    agg_dict = {label: (src, fn) for label, src, fn in PIVOT_METRICS}

    pivot = (
        df.groupby(GROUPBY_KEYS, dropna=False, observed=True)
        .agg(**agg_dict)
        .reset_index()
    )

    # 按 Zero_Bill 分类顺序排序
    pivot["Zero_Bill"] = pd.Categorical(pivot["Zero_Bill"], categories=ZERO_BILL_ORDER, ordered=True)
    pivot = pivot.sort_values(GROUPBY_KEYS).reset_index(drop=True)

    # 追加 Grand Total 行
    total_row = {"SCHEME_CODE": "Grand Total", "BILL_PERIOD": "", "Zero_Bill": "", "BILL_STATUS": ""}
    for label, src, fn in PIVOT_METRICS:
        total_row[label] = pivot[label].sum() if fn in ("sum", "count") else ""
    pivot = pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)

    _info(f"透视表构建完成：{len(pivot)} 行（含总计行）× {len(pivot.columns)} 列")
    return pivot


# ════════════════════════════════════════════════════════════════
# Step 6: 导出 xlsx（带格式）
# ════════════════════════════════════════════════════════════════
def export(pivot: pd.DataFrame, cfg: dict) -> Path:
    out_cfg = cfg.get("output", {})
    folder  = BASE_DIR / out_cfg.get("folder", "output")
    prefix  = out_cfg.get("filename_prefix", "pivot_result")
    sheet_n = out_cfg.get("sheet_name", "透视表")

    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = folder / f"{prefix}_{timestamp}.xlsx"

    _info(f"正在写出文件：{out_path.name} ...")

    # 先用 pandas 写入数据
    pivot.to_excel(out_path, sheet_name=sheet_n, index=False)

    # 再用 openpyxl 做格式美化
    wb = load_workbook(out_path)
    ws = wb[sheet_n]
    _format_sheet(ws, pivot)
    wb.save(out_path)

    _info(f"✓ 完成！输出文件：{out_path}")
    return out_path


def _format_sheet(ws, pivot: pd.DataFrame):
    # 颜色定义
    COLOR_HEADER    = "366092"   # 深蓝（表头背景）
    COLOR_TOTAL     = "D9E1F2"   # 浅蓝（总计行背景）
    COLOR_GROUP_L1  = "BDD7EE"   # 一级分组（SCHEME_CODE 变化行）
    COLOR_GROUP_L2  = "DDEBF7"   # 二级分组（BILL_PERIOD 变化行）
    COLOR_WHITE     = "FFFFFF"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font    = Font(name="Arial", size=10)
    total_font   = Font(name="Arial", bold=True, size=10)

    # 数字格式
    int_cols   = {"# of Pre-gen RS", "Sum of MB_CNT", "# of Submit RS",
                  "Sum of RS_SUBMIT_MEM", "# of Paid RS", "Sum of PAID_MEM"}
    money_cols = {"Bill Amount", "Submit Amount", "Sum of PAID_AMT", "Sum of OS_AMT"}

    headers = [c.value for c in ws[1]]

    # 表头行
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 30

    # 数据行
    total_row_idx = len(pivot)  # 最后一行是总计（1-indexed offset by header）
    prev_scheme = prev_period = None

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
        is_total = (r_idx == total_row_idx)
        scheme_val = row[0].value
        period_val = row[1].value

        # 行背景色
        if is_total:
            bg = COLOR_TOTAL
        elif scheme_val != prev_scheme:
            bg = COLOR_GROUP_L1
            prev_scheme = scheme_val
            prev_period = period_val
        elif period_val != prev_period:
            bg = COLOR_GROUP_L2
            prev_period = period_val
        else:
            bg = COLOR_WHITE

        for c_idx, cell in enumerate(row):
            col_name = headers[c_idx] if c_idx < len(headers) else ""
            cell.border = border
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.font   = total_font if is_total else body_font

            if col_name in int_cols and cell.value not in (None, ""):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col_name in money_cols and cell.value not in (None, ""):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # 自动列宽
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

    # 冻结首行
    ws.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════
# 工具函数
# ════════════════════════════════════════════════════════════════
def _info(msg: str):
    print(f"  [✓] {msg}")

def _warn(msg: str):
    print(f"  [!] {msg}")

def _abort(msg: str):
    print(f"\n  [✗] 错误：{msg}\n")
    sys.exit(1)


# ════════════════════════════════════════════════════════════════
# 主入口
# ════════════════════════════════════════════════════════════════
def main():
    print("\n" + "═" * 52)
    print("  透视表自动化脚本")
    print("═" * 52)

    cfg   = load_config()
    df    = load_data(cfg)
    df    = build_computed_cols(df)
    df_f  = apply_filters(df, cfg)
    pivot = build_pivot(df_f)
    out   = export(pivot, cfg)

    print("═" * 52)
    print(f"  全部完成！结果保存在：{out.relative_to(BASE_DIR)}")
    print("═" * 52 + "\n")


if __name__ == "__main__":
    main()
